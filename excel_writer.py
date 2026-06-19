import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import colorlog
import logging
from config import EXCEL_FILE, EXCEL_COLUMNS, LOG_FILE

# Cài đặt logger
handler = colorlog.StreamHandler()
handler.setFormatter(colorlog.ColoredFormatter(
    '%(log_color)s%(asctime)s - %(levelname)s - %(message)s'
))
file_handler = logging.FileHandler(LOG_FILE, encoding='utf-8')
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logger = colorlog.getLogger('excel_writer')
logger.setLevel(logging.DEBUG)
logger.addHandler(handler)
logger.addHandler(file_handler)

def init_excel_file():
    """Tạo file Excel mẫu nếu chưa tồn tại (BƯỚC 2)"""
    if os.path.exists(EXCEL_FILE):
        return
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TỔNG HỢP"
    
    # Định dạng header
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Ghi tiêu đề cột
    for col_num, column_title in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Căn chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 5   # STT
    ws.column_dimensions['B'].width = 20  # Số/Ký hiệu
    ws.column_dimensions['C'].width = 15  # Ngày VB
    ws.column_dimensions['D'].width = 25  # Cơ quan ban hành
    ws.column_dimensions['E'].width = 20  # Người ký
    ws.column_dimensions['F'].width = 40  # Trích yếu
    ws.column_dimensions['G'].width = 15  # Độ mật
    ws.column_dimensions['H'].width = 15  # Độ khẩn
    ws.column_dimensions['I'].width = 15  # Số đến
    ws.column_dimensions['J'].width = 15  # Ngày đến
    ws.column_dimensions['K'].width = 20  # Chủ trì
    ws.column_dimensions['L'].width = 20  # Ghi chú

    # Mock 5 dòng dữ liệu mẫu
    mock_data = [
        ["01-CV/TU", "01/06/2026", "Thành ủy TPHCM", "Nguyễn Văn A", "V/v tổ chức hội nghị", "Thường", "Thường", 1, "02/06/2026", "Nguyễn Hải Đăng", ""],
        ["02-QĐ/UBND", "02/06/2026", "UBND Phường", "Trần Văn B", "Quyết định bổ nhiệm", "Mật", "Khẩn", 2, "03/06/2026", "Tô Văn Sang", ""],
        ["03-TB/ĐU", "03/06/2026", "Đảng ủy Phường", "Lê Thị C", "Thông báo họp giao ban", "Thường", "Thường", 3, "03/06/2026", "Mai Thị Hoàng Phương", ""],
        ["04-KH/MTTQ", "04/06/2026", "MTTQ Phường", "Phạm Văn D", "Kế hoạch dân vận", "Thường", "Thường", 4, "04/06/2026", "Lâm Phong Thủy", ""],
        ["05-BC/KT", "05/06/2026", "UB Kiểm tra", "Hoàng Thị E", "Báo cáo giám sát", "Tối Mật", "Hỏa Tốc", 5, "05/06/2026", "Trần Duy Hào", ""]
    ]
    
    for row_num, row_data in enumerate(mock_data, 2):
        ws.cell(row=row_num, column=1, value=row_num-1)
        for col_num, value in enumerate(row_data, 2):
            ws.cell(row=row_num, column=col_num, value=value)
            
    # Lưu file
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    wb.save(EXCEL_FILE)
    logger.info("Đã tạo file Excel mẫu với 5 dòng mock data.")

def write_to_excel(doc_info, so_den):
    """
    Ghi dữ liệu văn bản vào file Excel (BƯỚC 3).
    - Mở file
    - Tìm dòng cuối
    - STT tự tăng
    - Ngày đến = hôm nay
    """
    try:
        init_excel_file()
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["TỔNG HỢP"]
        
        # Lấy số thứ tự tiếp theo
        last_row = ws.max_row
        next_stt = last_row
        
        # Ngày đến = hôm nay
        ngay_den = datetime.now().strftime("%d/%m/%Y")
        
        # Các cột: "STT", "Số/Ký hiệu", "Ngày văn bản", "Cơ quan ban hành", "Người ký",
        # "Trích yếu nội dung", "Độ mật", "Độ khẩn", "Số đến", "Ngày đến", "Chủ trì xử lý", "Ghi chú"
        
        row_data = [
            next_stt,
            doc_info.get("so_ky_hieu", ""),
            doc_info.get("ngay_van_ban", ""),
            doc_info.get("co_quan", ""),
            doc_info.get("nguoi_ky", ""),
            doc_info.get("trich_yeu", ""),
            doc_info.get("do_mat", ""),
            doc_info.get("do_khan", ""),
            so_den,
            ngay_den,
            doc_info.get("chu_tri", ""),
            "" # Ghi chú
        ]
        
        ws.append(row_data)
        wb.save(EXCEL_FILE)
        logger.info(f"Đã ghi vào Excel thành công: {doc_info.get('so_ky_hieu')}")
        return True
        
    except PermissionError:
        error_msg = f"LỖI: File Excel đang mở, không thể ghi! Vui lòng đóng file {EXCEL_FILE}"
        logger.error(error_msg)
        print("="*50)
        print(error_msg)
        print("="*50)
        return False
    except Exception as e:
        logger.error(f"Lỗi khi ghi Excel: {e}")
        return False
